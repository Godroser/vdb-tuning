#!/usr/bin/env python3
"""Calculate overlap probability between original-topN and sampled-topN rows."""

from __future__ import annotations

from pathlib import Path
from typing import Any

# ======================
# User-configurable params
# ======================
SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = SCRIPT_DIR / "sampling_ivf_param_sweep_results.xlsx"
TOP_N = 10
# Set to None to use active sheet.
SHEET_NAME: str | None = None
FITNESS_EPS = 1e-12


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).strip()
        if not text:
            return None
        return float(text)
    except Exception:
        return None


def row_id(row: dict[str, Any]) -> str:
    run_tag = row.get("run_tag")
    if run_tag:
        return f"run_tag={run_tag}"
    return (
        "sample_ratio={sample_ratio},M={M},ef_construction={ef_construction},"
        "ef_search={ef_search}"
    ).format(
        sample_ratio=row.get("sample_ratio"),
        M=row.get("M"),
        ef_construction=row.get("ef_construction"),
        ef_search=row.get("ef_search"),
    )


def top_n_ids(
    rows: list[dict[str, Any]],
    metric_col: str,
    top_n: int,
    *,
    higher_is_better: bool = True,
) -> set[str]:
    metric_rows: list[dict[str, Any]] = []
    for row in rows:
        metric = to_float(row.get(metric_col))
        if metric is not None:
            metric_rows.append({"id": row_id(row), "metric": metric})

    sorted_rows = sorted(
        metric_rows,
        key=lambda r: r["metric"],
        reverse=higher_is_better,
    )
    top_rows = sorted_rows[:top_n]
    return {r["id"] for r in top_rows}


def pareto_front_indices(points: list[tuple[float, float]]) -> set[int]:
    front: set[int] = set()
    for i, (qps_i, recall_i) in enumerate(points):
        dominated = False
        for j, (qps_j, recall_j) in enumerate(points):
            if i == j:
                continue
            if (
                qps_j >= qps_i
                and recall_j >= recall_i
                and (qps_j > qps_i or recall_j > recall_i)
            ):
                dominated = True
                break
        if not dominated:
            front.add(i)
    return front


def fitness_top_n_ids(
    rows: list[dict[str, Any]],
    qps_col: str,
    recall_col: str,
    top_n: int,
    eps: float = FITNESS_EPS,
) -> tuple[set[str], str | None, int]:
    valid_rows: list[dict[str, Any]] = []
    qps_values: list[float] = []
    recall_values: list[float] = []

    for row in rows:
        qps = to_float(row.get(qps_col))
        recall = to_float(row.get(recall_col))
        if qps is None or recall is None:
            continue
        valid_rows.append(row)
        qps_values.append(qps)
        recall_values.append(recall)

    if not valid_rows:
        return set(), None, 0

    max_qps = max(qps_values)
    max_recall = max(recall_values)
    if max_qps == 0:
        max_qps = 1.0
    if max_recall == 0:
        max_recall = 1.0

    points = list(zip(qps_values, recall_values))
    front_indices = pareto_front_indices(points)
    if not front_indices:
        return set(), None, 0

    ranked: list[tuple[str, float]] = []
    chosen_id: str | None = None
    best_fitness: float | None = None

    # Strictly rank only Pareto-front points.
    for i in sorted(front_indices):
        row = valid_rows[i]
        qps_norm = qps_values[i] / max_qps
        recall_norm = recall_values[i] / max_recall
        fitness = -1.0 / (abs(recall_norm - qps_norm) + eps)
        fitness = -fitness

        rid = row_id(row)
        ranked.append((rid, fitness))

        if best_fitness is None or fitness > best_fitness:
            best_fitness = fitness
            chosen_id = rid

    ranked.sort(key=lambda x: x[1], reverse=True)
    top_ids = {rid for rid, _fitness in ranked[:top_n]}
    return top_ids, chosen_id, len(front_indices)


def sample_ratio_key(value: Any) -> tuple[int, float | str]:
    ratio = to_float(value)
    if ratio is not None:
        return (0, ratio)
    return (1, str(value))


def main() -> None:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required to read xlsx. Install with: pip install openpyxl"
        ) from exc

    if TOP_N <= 0:
        raise ValueError("TOP_N must be a positive integer.")

    xlsx_path = XLSX_PATH.expanduser().resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        raise ValueError("Worksheet is empty.")

    header_names = [str(h).strip() if h is not None else "" for h in headers]
    required_cols = {
        "sample_ratio",
        "original_rps",
        "original_mean_precisions",
        "original_p95_time",
        "sampled_rps",
        "sampled_mean_precisions",
        "sampled_p95_time",
    }
    missing = [name for name in required_cols if name not in header_names]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    data_rows: list[dict[str, Any]] = []
    for values in rows_iter:
        row = {header_names[i]: values[i] for i in range(len(header_names))}
        data_rows.append(row)

    grouped_rows: dict[Any, list[dict[str, Any]]] = {}
    for row in data_rows:
        ratio = row.get("sample_ratio")
        grouped_rows.setdefault(ratio, []).append(row)

    if not grouped_rows:
        raise ValueError("No data rows found in worksheet.")

    print(f"xlsx: {xlsx_path}")
    print(f"sheet: {ws.title}")
    print(f"requested top_n: {TOP_N}")
    print(f"sample_ratio_groups: {len(grouped_rows)}")

    for ratio in sorted(grouped_rows.keys(), key=sample_ratio_key):
        rows = grouped_rows[ratio]
        original_rps_ids = top_n_ids(rows, "original_rps", TOP_N)
        sampled_rps_ids = top_n_ids(rows, "sampled_rps", TOP_N)
        original_prec_ids = top_n_ids(rows, "original_mean_precisions", TOP_N)
        sampled_prec_ids = top_n_ids(rows, "sampled_mean_precisions", TOP_N)
        original_p95_ids = top_n_ids(
            rows, "original_p95_time", TOP_N, higher_is_better=False
        )
        sampled_p95_ids = top_n_ids(
            rows, "sampled_p95_time", TOP_N, higher_is_better=False
        )
        original_fitness_ids, original_chosen, original_front_count = fitness_top_n_ids(
            rows=rows,
            qps_col="original_rps",
            recall_col="original_mean_precisions",
            top_n=TOP_N,
        )
        sampled_fitness_ids, sampled_chosen, sampled_front_count = fitness_top_n_ids(
            rows=rows,
            qps_col="sampled_rps",
            recall_col="sampled_mean_precisions",
            top_n=TOP_N,
        )

        print()
        print(f"===== sample_ratio = {ratio} =====")
        print(f"group_row_count: {len(rows)}")

        if original_rps_ids and sampled_rps_ids:
            overlap_rps_ids = sorted(original_rps_ids & sampled_rps_ids)
            rps_probability = len(overlap_rps_ids) / len(original_rps_ids)
            print("[RPS]")
            print(f"effective original_rps_top_n: {len(original_rps_ids)}")
            print(f"effective sampled_rps_top_n: {len(sampled_rps_ids)}")
            print(f"overlap_count: {len(overlap_rps_ids)}")
            print(
                f"overlap_probability: {rps_probability:.6f} "
                f"({len(overlap_rps_ids)}/{len(original_rps_ids)})"
            )
            # print("overlap rows (RPS):")
            # if overlap_rps_ids:
            #     for rid in overlap_rps_ids:
            #         print(f"- {rid}")
            # else:
            #     print("- (none)")
        else:
            print("[RPS]")
            print("- insufficient valid rows to compute top-N overlap.")

        if original_prec_ids and sampled_prec_ids:
            overlap_prec_ids = sorted(original_prec_ids & sampled_prec_ids)
            precision_probability = len(overlap_prec_ids) / len(original_prec_ids)
            print("[Mean Precisions]")
            print(f"effective original_mean_precisions_top_n: {len(original_prec_ids)}")
            print(f"effective sampled_mean_precisions_top_n: {len(sampled_prec_ids)}")
            print(f"overlap_count: {len(overlap_prec_ids)}")
            print(
                f"overlap_probability: {precision_probability:.6f} "
                f"({len(overlap_prec_ids)}/{len(original_prec_ids)})"
            )
            # print("overlap rows (Mean Precisions):")
            # if overlap_prec_ids:
            #     for rid in overlap_prec_ids:
            #         print(f"- {rid}")
            # else:
            #     print("- (none)")
        else:
            print("[Mean Precisions]")
            print("- insufficient valid rows to compute top-N overlap.")

        if original_p95_ids and sampled_p95_ids:
            overlap_p95_ids = sorted(original_p95_ids & sampled_p95_ids)
            p95_probability = len(overlap_p95_ids) / len(original_p95_ids)
            print("[P95 Time (lowest is better)]")
            print(f"effective original_p95_time_top_n: {len(original_p95_ids)}")
            print(f"effective sampled_p95_time_top_n: {len(sampled_p95_ids)}")
            print(f"overlap_count: {len(overlap_p95_ids)}")
            print(
                f"overlap_probability: {p95_probability:.6f} "
                f"({len(overlap_p95_ids)}/{len(original_p95_ids)})"
            )
        else:
            print("[P95 Time (lowest is better)]")
            print("- insufficient valid rows to compute top-N overlap.")

        if original_fitness_ids and sampled_fitness_ids:
            overlap_fitness_ids = sorted(original_fitness_ids & sampled_fitness_ids)
            fitness_probability = len(overlap_fitness_ids) / len(original_fitness_ids)
            print("[Pareto Fitness (RPS + Mean Precisions)]")
            print(f"original pareto_front_count: {original_front_count}")
            print(f"sampled pareto_front_count: {sampled_front_count}")
            print(f"effective original_fitness_top_n: {len(original_fitness_ids)}")
            print(f"effective sampled_fitness_top_n: {len(sampled_fitness_ids)}")
            print(f"original chosen point (argmax fitness): {original_chosen}")
            print(f"sampled chosen point (argmax fitness): {sampled_chosen}")
            print(f"overlap_count: {len(overlap_fitness_ids)}")
            print(
                f"overlap_probability: {fitness_probability:.6f} "
                f"({len(overlap_fitness_ids)}/{len(original_fitness_ids)})"
            )
            # print("overlap rows (Pareto Fitness):")
            # if overlap_fitness_ids:
            #     for rid in overlap_fitness_ids:
            #         print(f"- {rid}")
            # else:
            #     print("- (none)")
        else:
            print("[Pareto Fitness (RPS + Mean Precisions)]")
            print("- insufficient valid rows to compute top-N overlap.")


if __name__ == "__main__":
    main()
